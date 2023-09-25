from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog
import time
from  datetime import date
from datetime import *
from PIL import Image,ImageTk
import os
import openpyxl
import pathlib ,xlrd
from openpyxl import Workbook


ppro=Tk()

ppro.geometry('900x600+100+5')
ppro.configure(bg='#ffffff')
ppro.title('login')
ppro.resizable(False,False)


photo1=PhotoImage(file='login.png')
###############################lablel###
labpro=Label(ppro,image=photo1,bg='#ffffff').place(x=50,y=50)
###################################frame
fr=Frame(ppro,width=400,height=400,bg='#ffffff').place(x=450,y=50)
################################label
labfr1=Label(ppro,text='sign in ',fg='#2F98F1',bg='#ffffff',font='bold 20').place(x=600,y=50)

def exit ():
    ppro.destroy()
def signin ():
    username=user.get()
    password=code.get()
    pp=str({username:password})
    file=open('password\password.db','+r')
    ppp=file.read()
    file.close()
    if pp == ppp :
        messagebox.showinfo('welcome','Welcome back  !~'  +(username))
        ppro.destroy()
        student()
    else:
        messagebox.showerror('Error', 'try agin later')
          
#######################      
###############################################################

##############___________________ singup __________-________####

################################################################
def signupadmin():
    ppro.destroy()
    proo =Tk()
    proo.geometry('900x600+100+5')
    proo.configure(bg='#ffffff')
    proo.title('sign up')
    proo.resizable(False,False)


    photo1=PhotoImage(file='signup.png')
###############################lablel###
    labpro=Label(proo,image=photo1,bg='#ffffff')
    labpro.place(x=50,y=50)
###################################frame
    fr=Frame(proo,width=400,height=400,bg='#ffffff')
    fr.place(x=450,y=50)
################################label
    labfr1=Label(proo,text='sign up ',fg='#2F98F1',bg='#ffffff',font='bold 20').place(x=600,y=50)

    def exit():
        proo.destroy() 
#############################################
    def signup():
        username=user.get()
        password=code.get()
        conform=code1.get()
        if password == conform :
            file=open('password\password.db','w',)
            pp=str({username:password})
            file.write(pp)
            file.close()
        
            file=open('password\password.db','+r')
            ppp=file.read()
            file.close

            messagebox.showinfo('sign up','Successfully sing up ! '+(username))
            
            loginadmin()

        else:
            messagebox.showerror('Error,','try agin later') 
#######################################################
       

################################################
    def loginadmin():
        proo.destroy()
        proadmin=Tk()
        proadmin.geometry('900x600+100+5')
        proadmin.configure(bg='#ffffff')
        proadmin.title('login')
        proadmin.resizable(False,False)


        photo1=PhotoImage(file='login.png')
###############################lablel###
        labpro=Label(proadmin,image=photo1,bg='#ffffff').place(x=50,y=50)
###################################frame
        fr=Frame(proadmin,width=400,height=400,bg='#ffffff').place(x=450,y=50)
################################label
        labfr1=Label(fr,text='sign in ',fg='#2F98F1',bg='#ffffff',font='bold 20').place(x=600,y=50)

        def exit ():
            proadmin.destroy
        def signin ():
            username=user.get()
            password=code.get()
            pp=str({username:password})
            file=open('password\password.db','+r')
            ppp=file.read()
            file.close()
            if pp == ppp :
                messagebox.showinfo('welcome','Welcome back  !~'  +(username))
                proadmin.destroy()
                student()
            else:
                messagebox.showerror('Error', 'try agin later')
                  
#######################      
        def signup():
            pass



##########################################username
        def enter(e):
            user.delete(0,'end')
        def leave (e):
            name=user.get()
            if name=="":
                user.insert(0,'username')

        user=Entry(proadmin,width=20,fg='#000000',bd=0,bg='#ffffff',font='bold 15')
        user.place(x=550,y=100)
        user.insert(0,'username')
        Frame(proadmin,width=225,height=2,bg='#000000').place(x=550,y=125)
        user.bind('<FocusIn>',enter)
        user.bind('<FocusOut>',leave)
############################################# password
        def enter(e):
            code.delete(0,'end')
        def leave (e):
            name=code.get()
            if name=="":
                code.insert(0,'password')

        code=Entry(proadmin,width=20,fg='#000000',bd=0,bg='#ffffff',font='bold 15')
        code.place(x=550,y=200)
        code.insert(0,'password')
        Frame(proadmin,width=225,height=2,bg='#000000').place(x=550,y=225)
        code.bind('<FocusIn>',enter)
        code.bind('<FocusOut>',leave)
####################################################
        signinbutton=Button(proadmin,width=40,pady=7,text='sign in',bg='#2F98F1',fg='#ffffff',bd=0,command=signin).place(x=550,y=300,)
###################################
        lb1=Label(proadmin,text="Don't have an account ?",fg='#000000',bg='#ffffff',font='bold 9')
        lb1.place(x=550 ,y= 230)
        signupbutton=Button(proadmin,text='sign up now',bd=0,bg='#ffffff',cursor="hand2",width=10,fg='#2F98F1',command=signup)
        signupbutton.place(x=690 ,y= 230)
        proadmin.mainloop()

    


#############################################  ---user
    def enter(e):
        user.delete(0,'end')
    def leave (e):
        name=user.get()
        if name=="":
            user.insert(0,'username')

    user=Entry(proo,width=20,fg='#000000',bd=0,bg='#ffffff',font='bold 15')
    user.place(x=550,y=130)
    user.insert(0,'username')
    Frame(proo,width=225,height=2,bg='#000000').place(x=550,y=160)
    user.bind('<FocusIn>',enter)
    user.bind('<FocusOut>',leave)

#################################################  pass
    def enter(e):
        code.delete(0,'end')
    def leave (e):
        name=code.get()
        if name=="":
            code.insert(0,'password')


    code=Entry(proo,width=20,fg='#000000',bd=0,bg='#ffffff',font='bold 15')
    code.place(x=550,y=200)
    code.insert(0,'password')
    Frame(proo,width=225,height=2,bg='#000000').place(x=550,y=230)
    code.bind('<FocusIn>',enter)
    code.bind('<FocusOut>',leave)
######################################confirm pass
    def enter(e):
        code1.delete(0,'end')
    def leave (e):
        name=code1.get()
        if name=="":
            code1.insert(0,'conform password')

    code1=Entry(proo,width=20,fg='#000000',bd=0,bg='#ffffff',font='bold 15')
    code1.place(x=550,y=270)
    code1.insert(0,'conform password')
    Frame(proo,width=225,height=2,bg='#000000').place(x=550,y=300)
    code1.bind('<FocusIn>',enter)
    code1.bind('<FocusOut>',leave)






################################# button 
    signup1=Button(proo,pady=7,width=20,text='Registration',bg='#2F98F1',fg='#ffffff',bd=0,font='bold 16',command=signup).place(x=550,y=340,)
    lb1=Label(proo,text="I have an account ?",fg='#000000',bg='#ffffff',font='bold 9')
    lb1.place(x=550 ,y= 400)
    loginbutton=Button(proo,text='login',bd=0,bg='#ffffff',cursor="hand2",width=15,fg='#2F98F1',font='bold 10',command=loginadmin)
    loginbutton.place(x=690 ,y= 400)



    proo.mainloop()


##########################################username
def enter(e):
    user.delete(0,'end')
def leave (e):
    name=user.get()
    if name=="":
        user.insert(0,'username')

user=Entry(ppro,width=20,fg='#000000',bd=0,bg='#ffffff',font='bold 15')
user.place(x=550,y=100)
user.insert(0,'username')
Frame(ppro,width=225,height=2,bg='#000000').place(x=550,y=125)
user.bind('<FocusIn>',enter)
user.bind('<FocusOut>',leave)
############################################# password
def enter(e):
    code.delete(0,'end')
def leave (e):
    name=code.get()
    if name=="":
        code.insert(0,'password')

code=Entry(ppro,width=20,fg='#000000',bd=0,bg='#ffffff',font='bold 15')
code.place(x=550,y=200)
code.insert(0,'password')
Frame(ppro,width=225,height=2,bg='#000000').place(x=550,y=225)
code.bind('<FocusIn>',enter)
code.bind('<FocusOut>',leave)
####################################################
signinub=Button(fr,width=40,pady=7,text='sign in',bg='#2F98F1',fg='#ffffff',bd=0,command=signin).place(x=550,y=300,)
###################################
lb1=Label(ppro,text="Don't have an account ?",fg='#000000',bg='#ffffff',font='bold 9')
lb1.place(x=550 ,y= 230)
signupbu=Button(ppro,text='sign up now',bd=0,bg='#ffffff',cursor="hand2",width=10,fg='#2F98F1',command=signupadmin)
signupbu.place(x=690 ,y= 230)

###############################################################

#######------- student____________________###

##########################################################
def student():
    pro =Tk()
    pro.geometry('1200x650+100+10')
    pro.title('student')
    pro.config(bg= '#ffffff')
    pro.resizable(False,False)
#################################################

    file=pathlib.Path('Student_data.xlsx')
    if file.exists():
        pass
    else:    
        file = Workbook()
        sh=file.active
        sh['A1']='Name'
        sh['B1']='Age'
        sh['C1']='Phone'
        sh['D1']='Gmail'
        sh['E1']='F or M'
        sh['F1']='Number F or M'
        sh['G1']='Gender'
        sh['H1']='Payment'
        sh['I1']='Datetime'

        file.save('Student_data.xlsx')


#####################################################
    def Exit ():
        pro.destroy()


###################################
    search=StringVar()
    name=StringVar()
    age=StringVar()
    phone=StringVar()
    mail=StringVar()
    form=StringVar()
    number=StringVar()
    date=StringVar()
    gender=StringVar()
    payment=StringVar()


#############################################-------showimage
    def showimage():
        imagetxt=name.get()
        fileimage=filedialog.askopenfilename(initialdir=os.getcwd(),title='Select Image',filetypes=(('PNG file','PNG'),
                                                                                                ('JPG file','JPG'),
                                                                                                 ('All file','.txt')))


        img=(Image.open(fileimage))
        size=img.resize((150,150))
        phototk=ImageTk.PhotoImage(size)
        labimg.config(image=phototk)
        labimg.image=phototk
        img.save("images/" +str(imagetxt)+".png")

        messagebox.showinfo('upload',' Uploaded Successfully ! ' +(imagetxt))
 ##################################################################### clear
    def Restart ():
    
        name.set('')
        age.set('')   
        phone.set('')
        mail.set('')
        form.set('')
        number.set('')
        gender.set('Select')
        payment.set('Select')
        photo3=PhotoImage(file='im.png')
        labimg.config(image=photo3)
        labimg.image=photo3

# ###################################save
    def save():
        global img
        nameget=name.get()
        ageget=age.get()   
        phoneget=phone.get()
        mailget=mail.get()
        formget=form.get()
        numberget=number.get()
        genderget=gender.get()
        paymentget=payment.get()
        dateget=date.get()
        if nameget=='' or ageget=='' or phoneget == '' or mailget == '' or formget == '' or numberget == '' or genderget == '' or paymentget == '' or dateget =='' :
            messagebox.showerror('Error','NO Data ; try agin with set data')

        else:
            file=openpyxl.load_workbook('Student_data.xlsx')
            sheet=file.active
            sheet.cell(column=1 , row= sheet.max_row+1 ,value=nameget)   
            sheet.cell(column=2 , row= sheet.max_row ,value=ageget)
            sheet.cell(column=3 , row= sheet.max_row ,value=phoneget)
            sheet.cell(column=4 , row= sheet.max_row,value=mailget)
            sheet.cell(column=5 , row= sheet.max_row ,value=formget)
            sheet.cell(column=6 , row= sheet.max_row ,value=numberget)
            sheet.cell(column=7 , row= sheet.max_row ,value=genderget)
            sheet.cell(column=8 , row= sheet.max_row ,value=paymentget)
            sheet.cell(column=9 , row= sheet.max_row ,value=dateget)
            file.save(r'Student_data.xlsx')
        
            messagebox.showinfo('info','Successfully ')    

            Restart()
#################################### search 
    def searchnow():
        searchget=search.get()
        file=openpyxl.load_workbook('Student_data.xlsx')
        sheet=file.active

        for row in sheet.rows:

        
            if row[0].value == str (searchget):
                rowget=row[0]
                nu=str(rowget)[14:-1]
                numberrow=str(rowget)[15:-1]
#               print(nu)
#               print(number)
                m1=sheet.cell(row=int(numberrow),column=1).value
                m2=sheet.cell(row=int(numberrow),column=2).value
                m3=sheet.cell(row=int(numberrow),column=3).value
                m4=sheet.cell(row=int(numberrow),column=4).value
                m5=sheet.cell(row=int(numberrow),column=5).value
                m6=sheet.cell(row=int(numberrow),column=6).value
                m7=sheet.cell(row=int(numberrow),column=7).value
                m8=sheet.cell(row=int(numberrow),column=8).value
                m9=sheet.cell(row=int(numberrow),column=9).value

            #)
                
                name.set(m1)
                age.set(m2)   
                phone.set(m3)
                mail.set(m4)
                form.set(m5)
                number.set(m6)
                gender.set(m7)
                payment.set(m8)
                date.set(m9)
                img=(Image.open("images/" +str(m1)+".png"))
                size=img.resize((150,150))
                phototk=ImageTk.PhotoImage(size)
                labimg.config(image=phototk)
                labimg.image=phototk
    
    def updata():
        nameget=name.get()
        ageget=age.get()   
        phoneget=phone.get()
        mailget=mail.get()
        formget=form.get()
        numberget=number.get()
        genderget=gender.get()
        paymentget=payment.get()
        dateget=date.get()
        file=openpyxl.load_workbook('Student_data.xlsx')
        sheet=file.active

        for row in sheet.rows:

        
            if row[0].value == str (nameget):
                rowget=row[0]
                nu=str(rowget)[14:-1]
                numberrow=str(rowget)[15:-1]
        sheet.cell(column=1 , row= int(numberrow) ,value=nameget)   
        sheet.cell(column=2 , row= int(numberrow) ,value=ageget)
        sheet.cell(column=3 , row= int(numberrow) ,value=phoneget)
        sheet.cell(column=4 , row= int(numberrow),value=mailget)
        sheet.cell(column=5 , row= int(numberrow) ,value=formget)
        sheet.cell(column=6 , row= int(numberrow) ,value=numberget)
        sheet.cell(column=7 , row= int(numberrow) ,value=genderget)
        sheet.cell(column=8 , row= int(numberrow) ,value=paymentget)
        sheet.cell(column=9 , row= int(numberrow) ,value=dateget)
        file.save(r'Student_data.xlsx')
        if nameget =='' or ageget=='' or phoneget == '' or mailget == '' or formget == '' or numberget == '' or genderget == 'Select' or paymentget == 'Select' or dateget =='' :
            messagebox.showerror('Error','NO Data ; try agin with set data')
        else:      
            messagebox.showinfo('info','Successfully')
            Restart()    

            
        







#########################################################
    labpro=Label(pro,text='Welcome to Student Management Admin',bd=2,bg='#1d3557',fg='#ffffff',font='bold 20').pack(fill=X)
########################################################################################frame 1


    fr1=Frame(pro,width=1200,height=70,bg='#457b9d',bd=2).place(x=0,y=37)
    labfr1=Label(fr1,text='Student Search',bd=0,bg='#457b9d',fg='#ffffff',font='bold 20').place(x=600,y=70)
    enfr1=Entry(fr1,textvariable=search,font='Helvetica 10',bd=0,width=25,).place(x=800,y=70,height=30)
    bufr1=Button(fr1,text='Search',bg='#a8dadc',command=searchnow,fg='#ffffff',activeforeground='#000000',width=15,bd=0,height=2).place(x=1000,y=65)
  

   

###############################################frame 2
    fr2=Frame(pro,width=800,height=520,bg='#e63946',bd=2).place(x=20,y=120)
    labfr2=Label(fr2,text='ـــــــــــــــــــــ Student Detailsـــــــــــــــــــــ',fg='#ffffff',font='bold 20',bg='#e63946',bd=0).place(x=150,y=120)
    lab1=Label(fr2,text='Student Name',font='bold 15',bg='#e63946',fg='#ffffff',bd=0,).place(x=50,y=205)
    lab1=Label(fr2,text='Student Age',font='bold 15',bg='#e63946',fg='#ffffff',bd=0,).place(x=50,y=245)
    lab1=Label(fr2,text='Student phone',font='bold 15',bg='#e63946',fg='#ffffff',bd=0,).place(x=50,y=285)
    lab1=Label(fr2,text='Student Mail',font='bold 15',bg='#e63946',fg='#ffffff',bd=0,).place(x=50,y=325)
    lab1=Label(fr2,text='Father or Mother',font='bold 15',bg='#e63946',fg='#ffffff',bd=0,).place(x=50,y=365)
    lab1=Label(fr2,text='Phone Number (F,M)',font='bold 15',bg='#e63946',fg='#ffffff',bd=0,).place(x=50,y=405)
    lab1=Label(fr2,text='Datatime',font='bold 15',bg='#e63946',fg='#ffffff',bd=0,).place(x=50,y=445)
    lab1=Label(fr2,text='Gender ',font='bold 15',bg='#e63946',fg='#ffffff',bd=0,).place(x=100,y=500)
    lab1=Label(fr2,text='Installment Payment ',font='bold 15',bg='#e63946',fg='#ffffff',bd=0).place(x=450,y=500)
    combbox=ttk.Combobox(fr2,values=['Male','Female'],state='read',width=20,font='arial 13',textvariable=gender).place(x=100,y=525)
    gender.set('Select')
    combbox1=ttk.Combobox(fr2,values=['1000','2000','3000','4000','5000','6000','7000','8000','9000','10000',],textvariable=payment,state='read',width=20,font='bold 13').place(x=450,y=525)
    payment.set('Select')
    entname=Entry(fr2,width=20,fg='#000000',bd=0,bg='#ffffff',font='bold 15',textvariable=name).place(x=250,y=200)
    entage=Entry(fr2,width=20,fg='#000000',bd=0,bg='#ffffff',font='bold 15',textvariable=age).place(x=250,y=240)
    entphone=Entry(fr2,width=20,fg='#000000',bd=0,bg='#ffffff',font='bold 15',textvariable=phone).place(x=250,y=280)
    entmail=Entry(fr2,width=20,fg='#000000',bd=0,bg='#ffffff',font='bold 15',textvariable=mail).place(x=250,y=320)
    entfather=Entry(fr2,width=20,fg='#000000',bd=0,bg='#ffffff',font='bold 15',textvariable=form).place(x=250,y=360)
    entfnumber=Entry(fr2,width=20,fg='#000000',bd=0,bg='#ffffff',font='bold 15',textvariable=number).place(x=250,y=400)
    alltime=datetime.now()
    today=alltime.strftime('%d:%m:%y  |  %I:%M')
    entdate=Entry(fr2,width=20,fg='#000000',bd=0,bg='#ffffff',font='bold 15',textvariable=date).place(x=250,y=440)
    date.set(today)

###########################################################################3
    photo2=PhotoImage(file='im.png')
    labimg=Label(pro,bg='#ffffff',image=photo2,width=150,height=150)
    labimg.place(x=950,y=110)
    uploadimage=Button(pro,text='Upload image',bg='#1d3557',fg='#ffffff',activeforeground='#000000',command=showimage,font='bold 13',width=35,bd=0,height=2)
    uploadimage.place(x=860,y=270)
    savebutton=Button(pro,text='Save',command=save,bg='#457b9d',fg='#ffffff',activeforeground='#000000',font='bold 13',width=35,bd=0,height=2)
    savebutton.place(x=860,y=330)
    bupro1=Button(pro,text='UPdata',command=updata,bg='#a8dadc',fg='#ffffff',activeforeground='#000000',font='bold 13',width=35,bd=0,height=2).place(x=860,y=390)
    bupro1=Button(pro,text='Restart',command=Restart,bg='#2a9d8f',fg='#ffffff',activeforeground='#000000',font='bold 13',width=35,bd=0,height=2).place(x=860,y=450)
    bupro1=Button(pro,text=' Exit',bg='#e63946',command=Exit,fg='#ffffff',activeforeground='#000000',font='bold 13',width=35,bd=0,height=2).place(x=860,y=510)







    pro.mainloop()


ppro.mainloop()